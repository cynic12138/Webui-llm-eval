"""
Report generation service - PDF, Excel, JSON.
"""
import json
import io
from datetime import datetime, timezone
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.db import models
from app.services.storage import StorageService


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def generate(self, task: models.EvaluationTask, user: models.User, fmt: str) -> models.Report:
        if fmt == "pdf":
            content = await self._generate_pdf(task)
            ext = "pdf"
            ct = "application/pdf"
        elif fmt == "excel":
            content = await self._generate_excel(task)
            ext = "xlsx"
            ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            content = await self._generate_json(task)
            ext = "json"
            ct = "application/json"

        storage = StorageService()
        file_path = f"reports/{user.id}/{task.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
        await storage.upload_bytes(file_path, content, ct)

        report = models.Report(
            task_id=task.id,
            user_id=user.id,
            format=fmt,
            file_path=file_path,
            file_size=len(content),
        )
        self.db.add(report)
        await self.db.flush()
        await self.db.refresh(report)
        return report

    async def _generate_json(self, task: models.EvaluationTask) -> bytes:
        # Fetch results
        result = await self.db.execute(
            select(models.EvaluationResult).where(
                models.EvaluationResult.task_id == task.id
            ).limit(1000)
        )
        results = result.scalars().all()

        data = {
            "task_id": task.id,
            "task_name": task.name,
            "status": task.status,
            "created_at": task.created_at.isoformat(),
            "completed_at": task.completed_at.isoformat() if task.completed_at else None,
            "results_summary": task.results_summary,
            "sample_results": [
                {
                    "sample_index": r.sample_index,
                    "model_id": r.model_id,
                    "input": r.input_text,
                    "output": r.output_text,
                    "scores": r.scores,
                    "latency_ms": r.latency_ms,
                }
                for r in results
            ],
        }
        return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")

    async def _generate_excel(self, task: models.EvaluationTask) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary["A1"] = "LLM Evaluation Report"
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary["A2"] = f"Task: {task.name}"
        ws_summary["A3"] = f"Status: {task.status}"
        ws_summary["A4"] = f"Created: {task.created_at}"

        if task.results_summary:
            ws_summary["A6"] = "Model Scores"
            ws_summary["A6"].font = Font(bold=True)
            row = 7
            for model_id, model_data in task.results_summary.get("by_model", {}).items():
                ws_summary.cell(row=row, column=1, value=model_data.get("model_name", f"Model {model_id}"))
                col = 2
                for metric, score in model_data.get("scores", {}).items():
                    ws_summary.cell(row=6, column=col, value=metric).font = Font(bold=True)
                    ws_summary.cell(row=row, column=col, value=round(score, 4))
                    col += 1
                row += 1

        # Results sheet
        ws_results = wb.create_sheet("Results")
        result = await self.db.execute(
            select(models.EvaluationResult).where(
                models.EvaluationResult.task_id == task.id
            ).limit(5000)
        )
        results = result.scalars().all()

        headers = ["Sample", "Model ID", "Input", "Output", "Latency(ms)", "Prompt Tokens", "Completion Tokens"]
        for col, h in enumerate(headers, 1):
            ws_results.cell(row=1, column=col, value=h).font = Font(bold=True)

        for row_idx, r in enumerate(results, 2):
            ws_results.cell(row=row_idx, column=1, value=r.sample_index)
            ws_results.cell(row=row_idx, column=2, value=r.model_id)
            ws_results.cell(row=row_idx, column=3, value=(r.input_text or "")[:500])
            ws_results.cell(row=row_idx, column=4, value=(r.output_text or "")[:500])
            ws_results.cell(row=row_idx, column=5, value=r.latency_ms)
            ws_results.cell(row=row_idx, column=6, value=r.prompt_tokens)
            ws_results.cell(row=row_idx, column=7, value=r.completion_tokens)
            for col_offset, (key, val) in enumerate(r.scores.items(), 8):
                ws_results.cell(row=1, column=col_offset, value=key).font = Font(bold=True)
                ws_results.cell(row=row_idx, column=col_offset, value=round(float(val), 4) if isinstance(val, (int, float)) else str(val))

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def _generate_pdf(self, task: models.EvaluationTask) -> bytes:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("LLM Evaluation Report", styles["Title"]))
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"Task: {task.name}", styles["Heading2"]))
        story.append(Paragraph(f"Status: {task.status}", styles["Normal"]))
        story.append(Paragraph(f"Created: {task.created_at}", styles["Normal"]))
        story.append(Spacer(1, 12))

        if task.results_summary:
            story.append(Paragraph("Model Performance Summary", styles["Heading2"]))
            table_data = [["Model", "Metric", "Score"]]
            for model_id, model_data in task.results_summary.get("by_model", {}).items():
                model_name = model_data.get("model_name", f"Model {model_id}")
                for metric, score in model_data.get("scores", {}).items():
                    table_data.append([model_name, metric, f"{score:.4f}" if isinstance(score, float) else str(score)])
                table_data.append([model_name, "Avg Latency (ms)", f"{model_data.get('avg_latency_ms', 0):.1f}"])

            if len(table_data) > 1:
                t = Table(table_data)
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]))
                story.append(t)

        doc.build(story)
        return buf.getvalue()
